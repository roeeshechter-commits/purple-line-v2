"""
app.py — מערכת בקרת תעודות | הקו הסגול
Single-file Streamlit app — run with: streamlit run app.py
"""

# ═══════════════════════════════════════════════════════════════════════════════
#  IMPORTS
# ═══════════════════════════════════════════════════════════════════════════════

import io, json, re, csv, uuid, base64, zipfile, os
from datetime import datetime
from pathlib import Path

import streamlit as st
import pandas as pd

# ── Optional heavy imports (fail gracefully with instructions) ──────────────
try:
    import anthropic
    HAS_ANTHROPIC = True
except ImportError:
    HAS_ANTHROPIC = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import plotly.express as px
    import plotly.graph_objects as go
    HAS_PLOTLY = True
except ImportError:
    HAS_PLOTLY = False

try:
    from reportlab.lib import colors as rl_colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle,
        Paragraph, Spacer, HRFlowable,
    )
    from reportlab.lib.enums import TA_RIGHT, TA_CENTER
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False


# ═══════════════════════════════════════════════════════════════════════════════
#  PAGE CONFIG  (must be first Streamlit call)
# ═══════════════════════════════════════════════════════════════════════════════

st.set_page_config(
    page_title="בקרת תעודות — הקו הסגול",
    page_icon="🟣",
    layout="wide",
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Heebo:wght@300;400;500;600;700&display=swap');

html, body, [class*="css"]        { font-family: 'Heebo', sans-serif; direction: rtl; }
section[data-testid="stSidebar"]  { display: none; }

.top-bar {
    background: linear-gradient(135deg, #1e1b4b 0%, #312e81 60%, #4338ca 100%);
    padding: 1.4rem 2rem; border-radius: 14px; margin-bottom: 1.6rem;
    display: flex; align-items: center; justify-content: space-between;
}
.top-bar h1 { color: #fff; font-size: 1.55rem; font-weight: 700; margin: 0; }
.top-bar p  { color: rgba(255,255,255,.65); font-size: .85rem; margin: .2rem 0 0; }
.top-badge  {
    background: rgba(255,255,255,.12); color: #c7d2fe;
    padding: 4px 14px; border-radius: 20px; font-size: .78rem;
    border: 1px solid rgba(255,255,255,.2);
}

.kpi-grid { display: grid; grid-template-columns: repeat(4, 1fr); gap: 14px; margin-bottom: 1.4rem; }
.kpi-card {
    background: #fff; border: 1px solid #e5e7eb; border-radius: 12px;
    padding: 1.1rem 1.3rem; text-align: right;
}
.kpi-card .val  { font-size: 2.1rem; font-weight: 700; line-height: 1; }
.kpi-card .lbl  { font-size: .75rem; color: #6b7280; margin-top: 4px; }
.kpi-ok   .val  { color: #16a34a; }
.kpi-err  .val  { color: #dc2626; }
.kpi-warn .val  { color: #d97706; }
.kpi-info .val  { color: #4338ca; }

.quote-box {
    background: #faf5ff; border: 1px solid #e9d5ff;
    border-radius: 8px; padding: .9rem 1rem;
    font-family: monospace; font-size: .8rem;
    white-space: pre-wrap; direction: rtl; text-align: right;
    line-height: 1.7;
}
.issue-box {
    background: #fef2f2; border: 1px solid #fecaca;
    border-radius: 8px; padding: .8rem 1rem;
    font-size: .85rem; color: #991b1b;
}

/* Tab bar styling */
button[data-baseweb="tab"] { font-family: 'Heebo', sans-serif !important; font-size: .95rem !important; }
div[data-testid="stTabs"] > div > div { gap: 0 !important; }
</style>
""", unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════════════════════
#  CONSTANTS
# ═══════════════════════════════════════════════════════════════════════════════

HISTORY_FILE = Path("history.csv")

HISTORY_COLUMNS = [
    "batch_id", "timestamp", "month_year",
    "delivery_cert", "weighing_cert", "assmachta",
    "site", "driver", "vehicle", "date_on_doc", "net_weight_kg",
    "status",
    "check_delivery_in_excel", "check_weighing_in_excel", "check_assmachta_match",
    "issues", "quote_snippet", "source_file",
]


# ═══════════════════════════════════════════════════════════════════════════════
#  ██████╗  █████╗ ████████╗ █████╗     ██╗      █████╗ ██╗   ██╗███████╗██████╗
#  ██╔══██╗██╔══██╗╚══██╔══╝██╔══██╗    ██║     ██╔══██╗╚██╗ ██╔╝██╔════╝██╔══██╗
#  ██║  ██║███████║   ██║   ███████║    ██║     ███████║ ╚████╔╝ █████╗  ██████╔╝
#  ██║  ██║██╔══██║   ██║   ██╔══██║    ██║     ██╔══██║  ╚██╔╝  ██╔══╝  ██╔══██╗
#  ██████╔╝██║  ██║   ██║   ██║  ██║    ███████╗██║  ██║   ██║   ███████╗██║  ██║
#  ╚═════╝ ╚═╝  ╚═╝   ╚═╝   ╚═╝  ╚═╝    ╚══════╝╚═╝  ╚═╝   ╚═╝   ╚══════╝╚═╝  ╚═╝
# ═══════════════════════════════════════════════════════════════════════════════

# ── History helpers ──────────────────────────────────────────────────────────

def load_history() -> pd.DataFrame:
    if HISTORY_FILE.exists():
        df = pd.read_csv(HISTORY_FILE, dtype=str)
        for col in HISTORY_COLUMNS:
            if col not in df.columns:
                df[col] = ""
        return df[HISTORY_COLUMNS].fillna("")
    return pd.DataFrame(columns=HISTORY_COLUMNS)


def append_to_history(records: list[dict]) -> None:
    existing = load_history()
    new_rows  = pd.DataFrame(records, columns=HISTORY_COLUMNS)
    updated   = pd.concat([existing, new_rows], ignore_index=True)
    updated.to_csv(HISTORY_FILE, index=False, encoding="utf-8-sig")


def clear_history() -> None:
    if HISTORY_FILE.exists():
        HISTORY_FILE.unlink()


# ── Excel register parser ────────────────────────────────────────────────────

def parse_excel_register(excel_bytes: bytes) -> dict:
    """Read the חפורת Excel and return sets of known cert numbers."""
    if not HAS_OPENPYXL:
        st.error("openpyxl לא מותקן. הרץ: pip install openpyxl")
        return {"delivery_certs": set(), "weighing_certs": set(), "df": pd.DataFrame()}

    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), read_only=True, data_only=True)
    sheet_name = next(
        (s for s in wb.sheetnames if "חפורת" in s or "פסולת" in s),
        wb.sheetnames[0],
    )
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"delivery_certs": set(), "weighing_certs": set(), "df": pd.DataFrame()}

    # Locate header row
    header_idx = 0
    for i, row in enumerate(rows):
        joined = " ".join(str(c) for c in row if c)
        if "Delivery Note" in joined or "תעדות" in joined or "תעודת" in joined:
            header_idx = i
            break

    headers   = [str(c) if c is not None else f"col_{j}" for j, c in enumerate(rows[header_idx])]
    data_rows = rows[header_idx + 1:]
    df = pd.DataFrame(data_rows, columns=headers).dropna(how="all")

    delivery_col = headers[0]
    weighing_col = next(
        (h for h in headers if "Disposal Note" in h or "קליטה" in h),
        headers[11] if len(headers) > 11 else headers[-1],
    )

    def to_str_set(col):
        return {str(v).strip() for v in df[col].dropna()
                if str(v).strip() not in ("", "nan")}

    return {
        "delivery_certs": to_str_set(delivery_col),
        "weighing_certs": to_str_set(weighing_col),
        "df": df,
    }


# ── ZIP-as-PDF image extractor ───────────────────────────────────────────────

def extract_images_from_zip(file_bytes: bytes) -> list[tuple[str, bytes]]:
    """Extract sorted JPEGs from a ZIP file (the project's pseudo-PDF format)."""
    images = []
    try:
        with zipfile.ZipFile(io.BytesIO(file_bytes)) as zf:
            names = sorted(
                [n for n in zf.namelist() if n.lower().endswith((".jpeg", ".jpg", ".png"))],
                key=lambda x: int("".join(filter(str.isdigit, x.split(".")[0])) or "0"),
            )
            for name in names:
                images.append((name, zf.read(name)))
    except zipfile.BadZipFile:
        pass
    return images


# ── Anthropic AI helpers ─────────────────────────────────────────────────────

def _encode_b64(data: bytes) -> str:
    return base64.standard_b64encode(data).decode()


def extract_cert_data_from_page(
    client,
    image_bytes: bytes,
    doc_type: str,          # "delivery" | "weighing"
) -> dict:
    """Send one scanned page to Claude Vision → structured dict."""

    if doc_type == "delivery":
        prompt = """אתה מנתח תעודות משלוח לוגיסטיות.
חלץ את הנתונים הבאים ב-JSON בלבד (ללא טקסט אחר):
{
  "cert_number": "מספר תעודת המשלוח",
  "vehicle": "מספר הרכב",
  "driver": "שם הנהג",
  "date": "תאריך DD/MM/YYYY",
  "raw_text": "הטקסט המלא",
  "quote_snippet": "השורה המדויקת עם מספר התעודה"
}
שדה חסר → null."""
    else:
        prompt = """אתה מנתח תעודות שקילה/קליטה לוגיסטיות.
חלץ את הנתונים הבאים ב-JSON בלבד (ללא טקסט אחר):
{
  "cert_number": "מספר תעודת השקילה/קליטה",
  "assmachta": "מספר האסמכתא = מספר תעודת המשלוח שמופיע בשדה אסמכתא/מספר תעודת לקוח",
  "site": "שם אתר הפינוי",
  "date": "תאריך DD/MM/YYYY",
  "net_weight_kg": "משקל נטו בקילוגרמים",
  "raw_text": "הטקסט המלא",
  "quote_snippet": "המשפט המדויק המכיל את שדה האסמכתא, ציטוט מלא"
}
שדה חסר → null. שדה assmachta קריטי — בדוק בזהירות."""

    resp = client.messages.create(
        model="claude-opus-4-5",
        max_tokens=1024,
        messages=[{
            "role": "user",
            "content": [
                {"type": "image",
                 "source": {"type": "base64",
                            "media_type": "image/jpeg",
                            "data": _encode_b64(image_bytes)}},
                {"type": "text", "text": prompt},
            ],
        }],
    )
    raw = resp.content[0].text.strip()
    raw = re.sub(r"^```(?:json)?\s*", "", raw)
    raw = re.sub(r"\s*```$", "", raw)
    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        return {"raw_text": raw, "cert_number": None, "assmachta": None}


def cross_check(delivery_cert, weighing_cert, assmachta,
                excel_delivery: set, excel_weighing: set) -> dict:
    """Three-way cross-check. Returns status dict."""
    issues = []
    c1 = str(delivery_cert) in excel_delivery
    c2 = str(weighing_cert) in excel_weighing
    c3 = bool(assmachta) and str(assmachta) == str(delivery_cert)

    if not c1:
        issues.append(f"תעודת משלוח {delivery_cert} לא נמצאה בקובץ החפורת")
    if not c2:
        issues.append(f"תעודת קליטה {weighing_cert} לא נמצאה בקובץ החפורת")
    if not c3:
        issues.append(f"אסמכתא '{assmachta}' אינה תואמת למשלוח '{delivery_cert}'")

    return {
        "status": "ok" if not issues else "error",
        "check_delivery_in_excel": c1,
        "check_weighing_in_excel": c2,
        "check_assmachta_match": c3,
        "issues": issues,
    }


def ai_batch_summary(client, results: list[dict]) -> str:
    """Ask Claude for a management summary of the batch."""
    snippet = json.dumps(
        [{"delivery": r["delivery_cert"], "status": r["status"], "issues": r["issues"]}
         for r in results],
        ensure_ascii=False,
    )
    resp = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=600,
        messages=[{"role": "user", "content":
            f"תוצאות בדיקת הצלבה (JSON):\n{snippet}\n\n"
            "כתוב סיכום ניהולי קצר בעברית (3-4 משפטים): "
            "כמה נבדקו, שיעור הצלחה, פירוט חריגות אם יש, המלצה. "
            "ענה בעברית בלבד, ללא כותרות."}],
    )
    return resp.content[0].text.strip()


# ── Stats helpers ─────────────────────────────────────────────────────────────

def compute_stats(df: pd.DataFrame) -> dict:
    if df.empty:
        return {"total": 0, "ok": 0, "errors": 0, "ok_pct": 0.0, "batches": 0}
    total = len(df)
    ok    = (df["status"] == "ok").sum()
    return {
        "total":   int(total),
        "ok":      int(ok),
        "errors":  int(total - ok),
        "ok_pct":  round(ok / total * 100, 1),
        "batches": df["batch_id"].nunique(),
    }


def history_by_month(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=["month_year", "total", "ok", "errors", "ok_pct"])
    g = df.groupby("month_year").agg(
        total  = ("status", "count"),
        ok     = ("status", lambda s: (s == "ok").sum()),
        errors = ("status", lambda s: (s != "ok").sum()),
    ).reset_index()
    g["ok_pct"] = (g["ok"] / g["total"] * 100).round(1)
    return g


# ── PDF export ───────────────────────────────────────────────────────────────

def generate_pdf(df: pd.DataFrame, month_label: str = "") -> bytes | None:
    if not HAS_REPORTLAB:
        return None

    C_DARK   = rl_colors.HexColor("#1e1b4b")
    C_MID    = rl_colors.HexColor("#4338ca")
    C_LIGHT  = rl_colors.HexColor("#e0e7ff")
    C_GREEN  = rl_colors.HexColor("#16a34a")
    C_GREEN_L= rl_colors.HexColor("#f0fdf4")
    C_RED    = rl_colors.HexColor("#dc2626")
    C_RED_L  = rl_colors.HexColor("#fef2f2")
    C_GRAY_L = rl_colors.HexColor("#f9fafb")
    C_BORDER = rl_colors.HexColor("#e5e7eb")
    C_MUTED  = rl_colors.HexColor("#6b7280")

    buf  = io.BytesIO()
    doc  = SimpleDocTemplate(buf, pagesize=A4,
                             rightMargin=1.8*cm, leftMargin=1.8*cm,
                             topMargin=2*cm,     bottomMargin=2*cm)
    st_  = getSampleStyleSheet()
    def ps(name, **kw):
        return ParagraphStyle(name, parent=st_["Normal"], **kw)

    title_s   = ps("T", fontSize=17, fontName="Helvetica-Bold",
                   textColor=C_DARK, alignment=TA_RIGHT, spaceAfter=4)
    sub_s     = ps("S", fontSize=9,  textColor=C_MUTED,  alignment=TA_RIGHT)
    section_s = ps("H", fontSize=12, fontName="Helvetica-Bold",
                   textColor=C_DARK, alignment=TA_RIGHT, spaceBefore=12, spaceAfter=6)
    note_s    = ps("N", fontSize=8,  textColor=C_MUTED,  alignment=TA_RIGHT)

    label = month_label or datetime.now().strftime("%m/%Y")
    story = [
        Paragraph(f"דוח בקרת תעודות — {label}", title_s),
        Paragraph("מערכת ניהול חפורת ופסולת בנייה | הקו הסגול | סולל בונה", sub_s),
        Paragraph(f"הופק: {datetime.now().strftime('%d/%m/%Y %H:%M')}", sub_s),
        HRFlowable(width="100%", thickness=2, color=C_MID, spaceAfter=12),
    ]

    # KPI table
    total = len(df)
    ok    = (df["status"] == "ok").sum() if not df.empty else 0
    err   = total - ok
    pct   = round(ok / total * 100, 1) if total else 0

    kpi = Table(
        [["סה״כ תעודות", "תקינות", "חריגות", "שיעור תקינות"],
         [str(total), str(ok), str(err), f"{pct}%"]],
        colWidths=[4*cm]*4,
    )
    kpi.setStyle(TableStyle([
        ("BACKGROUND",  (0,0), (-1,0), C_DARK),
        ("TEXTCOLOR",   (0,0), (-1,0), rl_colors.white),
        ("FONTNAME",    (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",    (0,0), (-1,0), 10),
        ("FONTSIZE",    (0,1), (-1,1), 22),
        ("FONTNAME",    (0,1), (-1,1), "Helvetica-Bold"),
        ("TEXTCOLOR",   (1,1),  (1,1), C_GREEN),
        ("TEXTCOLOR",   (2,1),  (2,1), C_RED),
        ("ALIGN",       (0,0), (-1,-1), "CENTER"),
        ("VALIGN",      (0,0), (-1,-1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0,1), (-1,1), [C_GRAY_L]),
        ("GRID",        (0,0), (-1,-1), 0.5, C_BORDER),
        ("TOPPADDING",  (0,0), (-1,-1), 8),
        ("BOTTOMPADDING",(0,0),(-1,-1), 8),
    ]))
    story += [kpi, Spacer(1, 0.6*cm)]

    # Main table
    story.append(Paragraph("פירוט תעודות שנבדקו", section_s))
    cols = [
        ("delivery_cert",  "תעודת משלוח", 3.2*cm),
        ("weighing_cert",  "תעודת שקילה", 4*cm),
        ("assmachta",      "אסמכתא",      3*cm),
        ("site",           "אתר פינוי",   2.8*cm),
        ("driver",         "נהג",         2.2*cm),
        ("date_on_doc",    "תאריך",       2.2*cm),
        ("net_weight_kg",  "נטו (ק״ג)",   2*cm),
        ("status",         "סטטוס",       2*cm),
    ]
    hdr    = [[c[1] for c in cols]]
    widths = [c[2] for c in cols]
    tdata  = hdr[:]
    for _, row in df.iterrows():
        r = [str(row.get(c[0], "") or "") for c in cols]
        r[-1] = "✓ תקין" if r[-1] == "ok" else "✗ חריגה"
        tdata.append(r)

    main_t = Table(tdata, colWidths=widths, repeatRows=1)
    ts = [
        ("BACKGROUND",    (0,0), (-1,0), C_DARK),
        ("TEXTCOLOR",     (0,0), (-1,0), rl_colors.white),
        ("FONTNAME",      (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",      (0,0), (-1,-1), 8),
        ("ALIGN",         (0,0), (-1,-1), "CENTER"),
        ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
        ("GRID",          (0,0), (-1,-1), 0.4, C_BORDER),
        ("TOPPADDING",    (0,0), (-1,-1), 5),
        ("BOTTOMPADDING", (0,0), (-1,-1), 5),
        ("ROWBACKGROUNDS",(0,1), (-1,-1), [rl_colors.white, C_GRAY_L]),
    ]
    for i, row in enumerate(tdata[1:], start=1):
        if row[-1].startswith("✓"):
            ts += [("TEXTCOLOR", (-1,i), (-1,i), C_GREEN),
                   ("FONTNAME",  (-1,i), (-1,i), "Helvetica-Bold")]
        else:
            ts += [("BACKGROUND",(0,i), (-1,i), C_RED_L),
                   ("TEXTCOLOR", (-1,i), (-1,i), C_RED),
                   ("FONTNAME",  (-1,i), (-1,i), "Helvetica-Bold")]
    main_t.setStyle(TableStyle(ts))
    story.append(main_t)

    # Error detail
    errs = df[df["status"] != "ok"] if not df.empty else pd.DataFrame()
    if not errs.empty:
        story += [Spacer(1, 0.5*cm), Paragraph("פירוט חריגות", section_s)]
        for _, row in errs.iterrows():
            raw = row.get("issues", "") or ""
            story.append(Paragraph(
                f"• תעודה {row.get('delivery_cert','')}: {raw}", note_s))

    story += [
        Spacer(1, 0.8*cm),
        HRFlowable(width="100%", thickness=0.5, color=C_BORDER),
        Paragraph("דוח זה נוצר אוטומטית ע״י מערכת בקרת תעודות הקו הסגול", note_s),
    ]

    doc.build(story)
    return buf.getvalue()


# ── Excel export ──────────────────────────────────────────────────────────────

def generate_excel(df: pd.DataFrame, month_label: str = "") -> bytes | None:
    if not HAS_OPENPYXL:
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "תוצאות בדיקה"
    ws.sheet_view.rightToLeft = True

    h_fill  = PatternFill("solid", fgColor="1e1b4b")
    ok_fill = PatternFill("solid", fgColor="f0fdf4")
    er_fill = PatternFill("solid", fgColor="fef2f2")
    kp_fill = PatternFill("solid", fgColor="e0e7ff")
    h_font  = Font(bold=True, color="FFFFFF", size=11)
    ok_font = Font(bold=True, color="16a34a")
    er_font = Font(bold=True, color="dc2626")
    kp_font = Font(bold=True, color="1e1b4b", size=12)
    thin    = Border(*[Side(style="thin", color="e5e7eb")]*4)
    c_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    r_align = Alignment(horizontal="right",  vertical="center", wrap_text=True)

    label = month_label or datetime.now().strftime("%m/%Y")

    # Title
    ws.merge_cells("A1:L1")
    tc       = ws["A1"]
    tc.value = f"דוח בקרת תעודות — {label}"
    tc.font  = Font(bold=True, color="1e1b4b", size=14)
    tc.fill  = kp_fill
    tc.alignment = r_align
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:L2")
    sc       = ws["A2"]
    sc.value = f"הקו הסגול | הופק: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    sc.font  = Font(italic=True, color="6b7280", size=10)
    sc.alignment = r_align

    # KPIs row 4
    total = len(df)
    ok_n  = (df["status"] == "ok").sum() if not df.empty else 0
    er_n  = total - ok_n
    pct   = round(ok_n / total * 100, 1) if total else 0
    for col, val in zip("ABCDEFGH",
                        ["סה״כ", total, "תקינות", ok_n, "חריגות", er_n, "% תקינות", f"{pct}%"]):
        c = ws[f"{col}4"]
        c.value, c.fill, c.font, c.alignment, c.border = val, kp_fill, kp_font, c_align, thin
    ws.row_dimensions[4].height = 22

    # Header row 6
    col_defs = [
        ("delivery_cert",           "תעודת משלוח",    15),
        ("weighing_cert",           "תעודת שקילה",    18),
        ("assmachta",               "אסמכתא",         13),
        ("site",                    "אתר פינוי",       15),
        ("driver",                  "נהג",              11),
        ("date_on_doc",             "תאריך",           13),
        ("net_weight_kg",           "נטו (ק״ג)",       11),
        ("status",                  "סטטוס",           10),
        ("check_delivery_in_excel", "משלוח ב-Excel",   15),
        ("check_weighing_in_excel", "שקילה ב-Excel",   15),
        ("check_assmachta_match",   "אסמכתא תואמת",   15),
        ("issues",                  "פירוט חריגה",     35),
    ]
    for ci, (_, hdr, width) in enumerate(col_defs, 1):
        c = ws.cell(row=6, column=ci, value=hdr)
        c.font, c.fill, c.alignment, c.border = h_font, h_fill, c_align, thin
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[6].height = 22

    # Data rows
    for ri, (_, row) in enumerate(df.iterrows(), 7):
        is_ok = str(row.get("status","")) == "ok"
        rfill = ok_fill if is_ok else er_fill
        for ci, (key, _, _) in enumerate(col_defs, 1):
            val = row.get(key, "") or ""
            if key == "status":
                val = "✓ תקין" if is_ok else "✗ חריגה"
            elif key in ("check_delivery_in_excel","check_weighing_in_excel","check_assmachta_match"):
                val = "✓" if str(val).lower() in ("true","1","yes") else "✗"
            c = ws.cell(row=ri, column=ci, value=val)
            c.fill, c.border, c.alignment = rfill, thin, c_align
            if key == "status":
                c.font = ok_font if is_ok else er_font
        ws.row_dimensions[ri].height = 18

    ws.freeze_panes = "A7"
    ws.auto_filter.ref = f"A6:{get_column_letter(len(col_defs))}6"

    # Raw sheet
    ws2 = wb.create_sheet("נתונים גולמיים")
    ws2.sheet_view.rightToLeft = True
    if not df.empty:
        ws2.append(list(df.columns))
        for _, row in df.iterrows():
            ws2.append([str(v) if pd.notna(v) else "" for v in row])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ═══════════════════════════════════════════════════════════════════════════════
#  ██╗   ██╗██╗
#  ██║   ██║██║
#  ██║   ██║██║
#  ██║   ██║██║
#  ╚██████╔╝██║
#   ╚═════╝ ╚═╝
# ═══════════════════════════════════════════════════════════════════════════════

# ── Top bar ───────────────────────────────────────────────────────────────────
st.markdown("""
<div class="top-bar">
  <div>
    <h1>🟣 מערכת בקרת תעודות — הקו הסגול</h1>
    <p>ניהול ואימות תעודות חפורת ופסולת בנייה | PLW2 רבין / אלוף שדה | סולל בונה</p>
  </div>
  <span class="top-badge">PLW2 · דצמבר 2025</span>
</div>
""", unsafe_allow_html=True)

# ── Dependency check ──────────────────────────────────────────────────────────
missing = []
if not HAS_ANTHROPIC:   missing.append("`anthropic`")
if not HAS_OPENPYXL:    missing.append("`openpyxl`")
if not HAS_PLOTLY:      missing.append("`plotly`")
if not HAS_REPORTLAB:   missing.append("`reportlab`")
if missing:
    st.warning(
        f"⚠️ חסרות תלויות: {', '.join(missing)}. "
        f"הרץ: `pip install {' '.join(m.strip('`') for m in missing)}`",
        icon="📦",
    )

# ── Tabs ──────────────────────────────────────────────────────────────────────
tab_dash, tab_upload, tab_history, tab_analytics = st.tabs([
    "🏠  דשבורד",
    "📤  העלאה ועיבוד",
    "📋  היסטוריה",
    "📊  ניתוח וייצוא",
])


# ══════════════════════════════════════════════════
#  TAB 1 — DASHBOARD
# ══════════════════════════════════════════════════
with tab_dash:
    df_hist = load_history()
    stats   = compute_stats(df_hist)

    # KPI cards
    st.markdown(f"""
    <div class="kpi-grid">
      <div class="kpi-card kpi-info">
        <div class="val">{stats['total']:,}</div>
        <div class="lbl">סה״כ תעודות שנבדקו</div>
      </div>
      <div class="kpi-card kpi-ok">
        <div class="val">{stats['ok']:,}</div>
        <div class="lbl">✅ תקינות</div>
      </div>
      <div class="kpi-card kpi-err">
        <div class="val">{stats['errors']:,}</div>
        <div class="lbl">❌ חריגות אימות</div>
      </div>
      <div class="kpi-card kpi-warn">
        <div class="val">{stats['ok_pct']}%</div>
        <div class="lbl">שיעור תקינות</div>
      </div>
    </div>
    """, unsafe_allow_html=True)

    if df_hist.empty:
        st.info("📂 אין נתונים עדיין. עבור לטאב **העלאה ועיבוד** כדי להריץ בדיקה ראשונה.")
        # Demo charts
        if HAS_PLOTLY:
            st.markdown("#### תצוגה לדוגמה")
            months  = ["09/2025","10/2025","11/2025","12/2025"]
            demo_ok = [28,35,31,39]
            demo_er = [2,1,3,0]
            c1, c2 = st.columns(2)
            with c1:
                demo_df = pd.DataFrame({"חודש":months,"תקין":demo_ok,"חריגה":demo_er})
                fig = px.bar(demo_df, x="חודש", y=["תקין","חריגה"],
                             color_discrete_map={"תקין":"#16a34a","חריגה":"#dc2626"},
                             barmode="stack", title="כמות בדיקות לאורך זמן (דוגמה)")
                fig.update_layout(height=280, plot_bgcolor="white", paper_bgcolor="white",
                                  font_family="Heebo")
                st.plotly_chart(fig, use_container_width=True)
            with c2:
                pcts = [93.3,97.2,91.2,100.0]
                fig2 = go.Figure(go.Scatter(x=months,y=pcts,mode="lines+markers+text",
                                           text=[f"{p}%" for p in pcts],
                                           textposition="top center",
                                           line=dict(color="#4338ca",width=3),
                                           marker=dict(size=8)))
                fig2.add_hline(y=95,line_dash="dash",line_color="#dc2626",
                              annotation_text="יעד 95%",annotation_position="right")
                fig2.update_layout(height=280,title="% תקינות לאורך זמן (דוגמה)",
                                   plot_bgcolor="white",paper_bgcolor="white",
                                   yaxis=dict(range=[85,105]),font_family="Heebo")
                st.plotly_chart(fig2, use_container_width=True)
    else:
        monthly = history_by_month(df_hist)
        if HAS_PLOTLY and not monthly.empty:
            c1, c2 = st.columns(2)
            with c1:
                fig = px.bar(monthly, x="month_year", y=["ok","errors"],
                             color_discrete_map={"ok":"#16a34a","errors":"#dc2626"},
                             barmode="stack",
                             labels={"month_year":"חודש","value":"כמות","variable":"סטטוס"},
                             title="כמות בדיקות לאורך זמן")
                fig.update_layout(height=300, plot_bgcolor="white",
                                  paper_bgcolor="white", font_family="Heebo",
                                  margin=dict(t=40,b=40,l=10,r=10))
                st.plotly_chart(fig, use_container_width=True)
            with c2:
                fig2 = go.Figure()
                fig2.add_trace(go.Scatter(
                    x=monthly["month_year"], y=monthly["ok_pct"],
                    mode="lines+markers+text",
                    text=[f"{v}%" for v in monthly["ok_pct"]],
                    textposition="top center",
                    line=dict(color="#4338ca",width=3),
                    marker=dict(size=8,color="#4338ca"),
                    name="% תקינות",
                ))
                fig2.add_hline(y=95,line_dash="dash",line_color="#dc2626",
                              annotation_text="יעד 95%",annotation_position="right")
                fig2.update_layout(height=300,title="שיעור תקינות לאורך זמן",
                                   plot_bgcolor="white",paper_bgcolor="white",
                                   yaxis=dict(range=[0,110],title="%"),
                                   font_family="Heebo",
                                   margin=dict(t=40,b=40,l=10,r=10))
                st.plotly_chart(fig2, use_container_width=True)

        # Recent records
        st.subheader("10 רשומות אחרונות")
        recent = df_hist.tail(10)[[
            "batch_id","delivery_cert","weighing_cert","assmachta",
            "site","driver","date_on_doc","status"
        ]].copy()
        recent["status"] = recent["status"].map({"ok":"✅ תקין","error":"❌ חריגה"})
        st.dataframe(
            recent.rename(columns={
                "batch_id":"מזהה קבוצה","delivery_cert":"תעודת משלוח",
                "weighing_cert":"תעודת שקילה","assmachta":"אסמכתא",
                "site":"אתר","driver":"נהג","date_on_doc":"תאריך","status":"סטטוס",
            }),
            use_container_width=True, hide_index=True,
        )


# ══════════════════════════════════════════════════
#  TAB 2 — UPLOAD & PROCESS
# ══════════════════════════════════════════════════
with tab_upload:
    st.subheader("📤 העלאה ועיבוד אוטומטי")

    # API key
    with st.expander("🔑 Anthropic API Key", expanded="api_key" not in st.session_state):
        api_input = st.text_input(
            "מפתח API",
            type="password",
            value=st.session_state.get("api_key",""),
            placeholder="sk-ant-...",
            help="השג מ-console.anthropic.com",
            key="api_input",
        )
        if st.button("💾 שמור מפתח"):
            if api_input.strip():
                st.session_state["api_key"] = api_input.strip()
                st.success("✅ מפתח נשמר לסשן")
            else:
                st.warning("הכנס מפתח תקין")

    if not st.session_state.get("api_key"):
        st.warning("⚠️ הכנס API Key כדי להמשיך.")
        st.stop()

    if not HAS_ANTHROPIC:
        st.error("📦 התקן anthropic: `pip install anthropic`")
        st.stop()

    st.markdown("---")
    c1, c2, c3 = st.columns(3)
    with c1:
        st.markdown("##### 📄 תעודות משלוח")
        delivery_file = st.file_uploader("תעודות משלוח", type=["pdf","zip"],
                                          key="del_up", label_visibility="collapsed")
        if delivery_file:
            st.success(f"✅ {delivery_file.name}")
    with c2:
        st.markdown("##### ⚖️ תעודות שקילה/קליטה")
        weighing_file = st.file_uploader("תעודות שקילה", type=["pdf","zip"],
                                          key="wei_up", label_visibility="collapsed")
        if weighing_file:
            st.success(f"✅ {weighing_file.name}")
    with c3:
        st.markdown("##### 📊 קובץ חפורת (Excel)")
        excel_file = st.file_uploader("קובץ חפורת", type=["xlsx","xls"],
                                       key="exc_up", label_visibility="collapsed")
        if excel_file:
            st.success(f"✅ {excel_file.name}")

    month_label = st.text_input("תיוג חודש/שנה", value=datetime.now().strftime("%m/%Y"),
                                 max_chars=7, key="month_lbl")

    st.markdown("---")

    all_uploaded = delivery_file and weighing_file and excel_file
    if not all_uploaded:
        st.info("📎 העלה את שלושת הקבצים כדי להתחיל.")

    run_btn = st.button("🚀 הפעל בדיקת הצלבה", type="primary",
                         use_container_width=True, disabled=not all_uploaded)

    if run_btn and all_uploaded:
        batch_id = f"batch_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:6]}"

        # ── Parse Excel ─────────────────────────────────────────────────
        with st.status("⚙️ מעבד קובץ חפורת Excel...", expanded=True) as s:
            try:
                register = parse_excel_register(excel_file.read())
                n_del = len(register["delivery_certs"])
                n_wei = len(register["weighing_certs"])
                st.write(f"✅ {n_del} תעודות משלוח + {n_wei} תעודות קליטה נקראו")
                s.update(label="✅ קובץ חפורת עובד", state="complete")
            except Exception as e:
                s.update(label=f"❌ שגיאה: {e}", state="error")
                st.error(str(e)); st.stop()

        # ── Extract images ───────────────────────────────────────────────
        with st.status("📸 חולץ תמונות...", expanded=True) as s:
            del_images = extract_images_from_zip(delivery_file.read())
            wei_images = extract_images_from_zip(weighing_file.read())
            if not del_images or not wei_images:
                s.update(label="❌ לא נמצאו תמונות", state="error")
                st.error("ודא שהקבצים הם ZIP המכילים JPEGs ממוספרים (1.jpeg, 2.jpeg...)")
                st.stop()
            st.write(f"✅ {len(del_images)} תמונות משלוח | {len(wei_images)} תמונות שקילה")
            s.update(label="✅ תמונות חולצו", state="complete")

        # ── AI extraction ────────────────────────────────────────────────
        client    = anthropic.Anthropic(api_key=st.session_state["api_key"])
        total_pg  = len(del_images) + len(wei_images)
        prog      = st.progress(0, text="מנתח תעודות עם Claude Vision...")
        done_pg   = 0
        del_data, wei_data = [], []

        st.write("**🤖 מנתח תעודות משלוח...**")
        for fname, img_bytes in del_images:
            try:
                d = extract_cert_data_from_page(client, img_bytes, "delivery")
            except Exception as e:
                d = {"cert_number": None, "raw_text": str(e)}
            d["_src"] = fname
            del_data.append(d)
            done_pg += 1
            prog.progress(done_pg / total_pg, text=f"עמוד {done_pg}/{total_pg}")

        st.write("**🤖 מנתח תעודות שקילה...**")
        for fname, img_bytes in wei_images:
            try:
                w = extract_cert_data_from_page(client, img_bytes, "weighing")
            except Exception as e:
                w = {"cert_number": None, "assmachta": None, "raw_text": str(e)}
            w["_src"] = fname
            wei_data.append(w)
            done_pg += 1
            prog.progress(done_pg / total_pg, text=f"עמוד {done_pg}/{total_pg}")

        prog.empty()

        # ── Cross-check + build records ──────────────────────────────────
        results = []
        with st.spinner("מבצע הצלבה..."):
            n_pairs = min(len(del_data), len(wei_data))
            for i in range(n_pairs):
                d = del_data[i]
                w = wei_data[i]
                dc = str(d.get("cert_number") or "")
                wc = str(w.get("cert_number") or "")
                am = str(w.get("assmachta")   or "")
                chk = cross_check(dc, wc, am,
                                  register["delivery_certs"],
                                  register["weighing_certs"])
                results.append({
                    "batch_id":                batch_id,
                    "timestamp":               datetime.now().isoformat(),
                    "month_year":              month_label,
                    "delivery_cert":           dc,
                    "weighing_cert":           wc,
                    "assmachta":               am,
                    "site":                    w.get("site") or "",
                    "driver":                  d.get("driver") or "",
                    "vehicle":                 d.get("vehicle") or "",
                    "date_on_doc":             d.get("date") or w.get("date") or "",
                    "net_weight_kg":           w.get("net_weight_kg") or "",
                    "status":                  chk["status"],
                    "check_delivery_in_excel": chk["check_delivery_in_excel"],
                    "check_weighing_in_excel": chk["check_weighing_in_excel"],
                    "check_assmachta_match":   chk["check_assmachta_match"],
                    "issues":                  json.dumps(chk["issues"], ensure_ascii=False),
                    "quote_snippet":           w.get("quote_snippet") or "",
                    "source_file":             d.get("_src") or "",
                })

        # ── Save + display ───────────────────────────────────────────────
        append_to_history(results)
        res_df  = pd.DataFrame(results)
        ok_n    = (res_df["status"] == "ok").sum()
        err_n   = (res_df["status"] != "ok").sum()

        st.success(f"✅ הבדיקה הושלמה! {ok_n} תקינות · {err_n} חריגות — נשמר לקובץ history.csv")

        with st.spinner("מחולל סיכום ניהולי..."):
            try:
                summary = ai_batch_summary(client, results)
                st.info(f"**📋 סיכום ניהולי:**\n\n{summary}")
            except Exception:
                pass

        # Results table
        disp = res_df[["delivery_cert","weighing_cert","assmachta",
                        "site","driver","date_on_doc","status","issues"]].copy()
        disp["status"] = disp["status"].map({"ok":"✅ תקין","error":"❌ חריגה"})
        st.dataframe(
            disp.rename(columns={
                "delivery_cert":"תעודת משלוח","weighing_cert":"תעודת שקילה",
                "assmachta":"אסמכתא","site":"אתר","driver":"נהג",
                "date_on_doc":"תאריך","status":"סטטוס","issues":"פירוט",
            }),
            use_container_width=True, hide_index=True,
        )

        # Error detail
        if err_n > 0:
            st.error(f"⚠️ {err_n} חריגות נמצאו")
            with st.expander("🔍 פירוט מלא של החריגות"):
                for r in results:
                    if r["status"] != "ok":
                        st.markdown(f"**תעודת משלוח {r['delivery_cert']}**")
                        issues_list = json.loads(r["issues"]) if r["issues"] else []
                        for iss in issues_list:
                            st.markdown(f"- {iss}")
                        if r.get("quote_snippet"):
                            st.markdown(
                                f"<div class='quote-box'>{r['quote_snippet']}</div>",
                                unsafe_allow_html=True,
                            )
                        st.markdown("---")

        st.session_state["last_results"] = results
        st.balloons()


# ══════════════════════════════════════════════════
#  TAB 3 — HISTORY
# ══════════════════════════════════════════════════
with tab_history:
    st.subheader("📋 היסטוריית בדיקות")

    df_h = load_history()
    if df_h.empty:
        st.info("📂 אין נתונים. הפעל בדיקה בטאב **העלאה ועיבוד**.")
    else:
        # Filter row
        fc1, fc2, fc3, fc4 = st.columns([2,2,2,3])
        with fc1:
            status_f = st.selectbox("סטטוס", ["הכל","✅ תקין","❌ חריגה"], key="hf_status")
        with fc2:
            months_opts = ["הכל"] + sorted(df_h["month_year"].unique().tolist(), reverse=True)
            month_f = st.selectbox("חודש", months_opts, key="hf_month")
        with fc3:
            sites_opts = ["הכל"] + sorted(df_h["site"].dropna().unique().tolist())
            site_f = st.selectbox("אתר פינוי", sites_opts, key="hf_site")
        with fc4:
            search_f = st.text_input("🔍 חיפוש חופשי", placeholder="מספר תעודה / נהג...", key="hf_search")

        # Apply
        fdf = df_h.copy()
        if status_f == "✅ תקין":   fdf = fdf[fdf["status"] == "ok"]
        elif status_f == "❌ חריגה": fdf = fdf[fdf["status"] != "ok"]
        if month_f != "הכל": fdf = fdf[fdf["month_year"] == month_f]
        if site_f  != "הכל": fdf = fdf[fdf["site"] == site_f]
        if search_f.strip():
            q = search_f.strip().lower()
            mask = (fdf["delivery_cert"].str.lower().str.contains(q, na=False) |
                    fdf["weighing_cert"].str.lower().str.contains(q, na=False) |
                    fdf["driver"].str.lower().str.contains(q, na=False))
            fdf = fdf[mask]

        # Mini KPIs
        m1,m2,m3,m4 = st.columns(4)
        m1.metric("תוצאות", len(fdf))
        m2.metric("תקינות", (fdf["status"]=="ok").sum())
        m3.metric("חריגות", (fdf["status"]!="ok").sum())
        pct_f = round((fdf["status"]=="ok").sum() / max(len(fdf),1)*100,1)
        m4.metric("שיעור תקינות", f"{pct_f}%")

        # Table
        disp_h = fdf[[
            "batch_id","month_year","delivery_cert","weighing_cert",
            "assmachta","site","driver","date_on_doc","net_weight_kg","status"
        ]].copy()
        disp_h["status"] = disp_h["status"].map({"ok":"✅ תקין","error":"❌ חריגה"})
        disp_h = disp_h.rename(columns={
            "batch_id":"מזהה","month_year":"חודש","delivery_cert":"תעודת משלוח",
            "weighing_cert":"תעודת שקילה","assmachta":"אסמכתא","site":"אתר",
            "driver":"נהג","date_on_doc":"תאריך","net_weight_kg":"נטו (ק״ג)","status":"סטטוס",
        })
        st.dataframe(disp_h, use_container_width=True, hide_index=True, height=380)

        # Row detail
        st.markdown("---")
        st.markdown("#### 🔍 פרטי שורה")
        row_n = st.number_input("מספר שורה לצפייה", min_value=1,
                                 max_value=max(len(fdf),1), value=1, step=1)
        if not fdf.empty:
            sel = fdf.iloc[row_n - 1]
            dc1, dc2 = st.columns(2)
            with dc1:
                st.markdown("**פרטי תעודה**")
                for k,v in [("תעודת משלוח",sel.get("delivery_cert","")),
                            ("תעודת שקילה", sel.get("weighing_cert","")),
                            ("אסמכתא",      sel.get("assmachta","")),
                            ("אתר פינוי",   sel.get("site","")),
                            ("נהג",         sel.get("driver","")),
                            ("מס׳ רכב",    sel.get("vehicle","")),
                            ("תאריך",       sel.get("date_on_doc","")),
                            ("נטו (ק״ג)",   sel.get("net_weight_kg",""))]:
                    st.markdown(f"**{k}:** {v}")
            with dc2:
                is_ok = str(sel.get("status","")) == "ok"
                if is_ok:
                    st.success("✅ תעודה תקינה")
                else:
                    st.error("❌ חריגת אימות")
                st.markdown("**בדיקות הצלבה:**")
                for label_c, key_c in [
                    ("תעודת משלוח בקובץ החפורת","check_delivery_in_excel"),
                    ("תעודת שקילה בקובץ החפורת","check_weighing_in_excel"),
                    ("אסמכתא תואמת למשלוח",     "check_assmachta_match"),
                ]:
                    v = str(sel.get(key_c,"")).lower()
                    ic = "✅" if v in ("true","1") else "❌"
                    st.markdown(f"{ic} {label_c}")

                raw_issues = sel.get("issues","") or ""
                if raw_issues and raw_issues != "[]":
                    try:    iss_list = json.loads(raw_issues)
                    except: iss_list = [raw_issues]
                    if iss_list:
                        st.markdown("**פירוט חריגות:**")
                        for iss in iss_list:
                            st.markdown(f"- {iss}")

            quote = sel.get("quote_snippet","") or ""
            if quote.strip() and quote != "nan":
                st.markdown("**ציטוט מהמסמך:**")
                st.markdown(f"<div class='quote-box'>{quote}</div>",
                            unsafe_allow_html=True)

        # Delete history
        st.markdown("---")
        with st.expander("🗑️ מחיקת היסטוריה"):
            st.warning("פעולה זו תמחק את כל קובץ history.csv ולא ניתן לשחזרה.")
            if st.button("מחק היסטוריה", type="secondary"):
                clear_history()
                st.success("ההיסטוריה נמחקה.")
                st.rerun()


# ══════════════════════════════════════════════════
#  TAB 4 — ANALYTICS & EXPORT
# ══════════════════════════════════════════════════
with tab_analytics:
    st.subheader("📊 ניתוח וייצוא")

    df_a = load_history()
    if df_a.empty:
        st.info("📂 אין נתונים לניתוח.")
    else:
        months_a   = sorted(df_a["month_year"].dropna().unique().tolist(), reverse=True)
        sel_month  = st.selectbox("בחר חודש לדוח", ["כל הזמנים"] + months_a, key="ana_month")
        df_fil     = df_a if sel_month == "כל הזמנים" else df_a[df_a["month_year"] == sel_month]
        month_lbl  = sel_month if sel_month != "כל הזמנים" else datetime.now().strftime("%m/%Y")

        if HAS_PLOTLY:
            # Site breakdown
            st.markdown("#### ניתוח לפי אתר פינוי")
            a1, a2 = st.columns(2)
            with a1:
                sc = df_fil.groupby(["site","status"]).size().reset_index(name="count")
                if not sc.empty:
                    fig = px.bar(sc, x="site", y="count", color="status",
                                 color_discrete_map={"ok":"#16a34a","error":"#dc2626"},
                                 barmode="group",
                                 labels={"site":"אתר","count":"כמות","status":"סטטוס"})
                    fig.update_layout(height=280, plot_bgcolor="white",
                                      paper_bgcolor="white", font_family="Heebo",
                                      margin=dict(t=20,b=60,l=10,r=10))
                    st.plotly_chart(fig, use_container_width=True)
            with a2:
                sp = df_fil["status"].value_counts()
                if not sp.empty:
                    fig2 = go.Figure(go.Pie(
                        labels=["✅ תקין" if l=="ok" else "❌ חריגה" for l in sp.index],
                        values=sp.values,
                        marker_colors=["#16a34a" if l=="ok" else "#dc2626" for l in sp.index],
                        hole=0.42, textinfo="label+percent",
                    ))
                    fig2.update_layout(height=280, paper_bgcolor="white",
                                       margin=dict(t=20,b=20,l=20,r=20))
                    st.plotly_chart(fig2, use_container_width=True)

            # Driver breakdown
            st.markdown("#### ניתוח לפי נהג")
            if "driver" in df_fil.columns and not df_fil.empty:
                drv = (df_fil.groupby("driver")
                       .agg(total=("status","count"),
                            ok   =("status", lambda s: (s=="ok").sum()),
                            errors=("status",lambda s: (s!="ok").sum()))
                       .reset_index())
                drv["ok_pct"] = (drv["ok"] / drv["total"] * 100).round(1)
                b1, b2 = st.columns([2,1])
                with b1:
                    fig3 = px.bar(drv, x="driver", y=["ok","errors"],
                                  color_discrete_map={"ok":"#16a34a","errors":"#dc2626"},
                                  barmode="stack",
                                  labels={"driver":"נהג","value":"כמות","variable":"סטטוס"})
                    fig3.update_layout(height=260, plot_bgcolor="white",
                                       paper_bgcolor="white", font_family="Heebo",
                                       margin=dict(t=20,b=60,l=10,r=10))
                    st.plotly_chart(fig3, use_container_width=True)
                with b2:
                    st.dataframe(drv.rename(columns={
                        "driver":"נהג","total":"סה״כ","ok":"תקין",
                        "errors":"חריגות","ok_pct":"% תקינות"}),
                        hide_index=True, height=220)

        # Monthly trend
        monthly_a = history_by_month(df_a)
        if HAS_PLOTLY and len(monthly_a) > 1:
            st.markdown("#### מגמה חודשית")
            fig4 = go.Figure()
            fig4.add_trace(go.Bar(x=monthly_a["month_year"], y=monthly_a["total"],
                                  name="סה״כ", marker_color="#c7d2fe"))
            fig4.add_trace(go.Scatter(x=monthly_a["month_year"], y=monthly_a["ok_pct"],
                                      mode="lines+markers", name="% תקינות",
                                      yaxis="y2",
                                      line=dict(color="#4338ca",width=3),
                                      marker=dict(size=7)))
            fig4.update_layout(
                plot_bgcolor="white", paper_bgcolor="white", height=290,
                font_family="Heebo",
                yaxis=dict(title="כמות תעודות"),
                yaxis2=dict(title="% תקינות", overlaying="y", side="left", range=[0,110]),
                legend=dict(orientation="h", y=1.15),
                margin=dict(t=20,b=40,l=10,r=10),
            )
            st.plotly_chart(fig4, use_container_width=True)

        # ── Export buttons ───────────────────────────────────────────────
        st.markdown("---")
        st.markdown("#### 📥 ייצוא דוח")
        e1, e2, e3 = st.columns(3)

        with e1:
            st.markdown("**📄 PDF**")
            if st.button("צור דוח PDF", use_container_width=True, key="pdf_btn"):
                with st.spinner("מייצר PDF..."):
                    if not HAS_REPORTLAB:
                        st.error("התקן reportlab: `pip install reportlab`")
                    else:
                        pdf_b = generate_pdf(df_fil, month_lbl)
                        if pdf_b:
                            st.download_button(
                                "⬇️ הורד PDF", pdf_b,
                                file_name=f"report_{month_lbl.replace('/','_')}.pdf",
                                mime="application/pdf",
                                use_container_width=True, key="pdf_dl",
                            )

        with e2:
            st.markdown("**📊 Excel**")
            if st.button("צור דוח Excel", use_container_width=True, key="xl_btn"):
                with st.spinner("מייצר Excel..."):
                    xl_b = generate_excel(df_fil, month_lbl)
                    if xl_b:
                        st.download_button(
                            "⬇️ הורד Excel", xl_b,
                            file_name=f"report_{month_lbl.replace('/','_')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True, key="xl_dl",
                        )
                    elif not HAS_OPENPYXL:
                        st.error("התקן openpyxl: `pip install openpyxl`")

        with e3:
            st.markdown("**📂 CSV גולמי**")
            csv_b = df_fil.to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig")
            st.download_button(
                "⬇️ הורד CSV", csv_b,
                file_name=f"history_{month_lbl.replace('/','_')}.csv",
                mime="text/csv",
                use_container_width=True, key="csv_dl",
            )
